from __future__ import annotations

import json
from collections import Counter, defaultdict
from datetime import date, datetime
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
SOURCE_DIR = ROOT / "Files"
OUTPUT_PATH = ROOT / "lib" / "generated" / "mock-source-data.json"
INDEX_OUTPUT_PATH = ROOT / "lib" / "generated" / "mock-course-index.json"
WORKING_COPY_DIR = ROOT / "tmp" / "lms-import"

LMS_FILES = [
    "all_Single Video Courses.xlsx",
    "all_Standard Courses.xlsx",
    "all_Training Block Courses.xlsx",
    "all_Full Length Courses.xlsx",
]

CANONICAL_LMS_HEADERS = {
    "id": "Course ID",
    "name": "Course Name",
    "content_type": "Course Type",
    "training_credits": "Training Credits",
    "published": "Is published",
    "issuing_body": "Issuing Body",
    "state": "State",
    "accreditation_number": "Accreditation Number",
    "topic_number": "Topic Number",
    "start_date": "Accreditation Start Date",
    "end_date": "Accreditation End Date",
}


def json_value(value):
    if isinstance(value, (datetime, date)):
        return value.date().isoformat() if isinstance(value, datetime) else value.isoformat()
    if isinstance(value, float) and value.is_integer():
        return int(value)
    return value


def read_records(path: Path):
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook.active
    rows = sheet.iter_rows(values_only=True)
    headers = [str(value).replace("\u00a0", " ").strip() if value is not None else "" for value in next(rows)]
    records = []
    for values in rows:
        record = {
            headers[index]: json_value(value)
            for index, value in enumerate(values)
            if index < len(headers) and headers[index]
        }
        if any(value not in (None, "") for value in record.values()):
            records.append(record)
    workbook.close()
    return headers, records


def course_id(record):
    value = record.get("Course ID", record.get("Course Id", record.get("id")))
    if value in (None, ""):
        return None
    if isinstance(value, float) and value.is_integer():
        value = int(value)
    return str(value).strip()


def canonical_lms_record(record):
    canonical = {}
    for key, value in record.items():
        output_key = CANONICAL_LMS_HEADERS.get(key, key)
        canonical[output_key] = value
    return canonical


def normalize_verticals(value):
    if value in (None, ""):
        return None
    normalized = str(value).replace("EMS1A", "EMS1")
    return normalized


def topic_vertical(column_index):
    if column_index <= 16:
        return "LGU"
    if column_index <= 21:
        return "Lexipol"
    if column_index <= 44:
        return "FR1A"
    if column_index <= 64:
        return "EMS1"
    if column_index <= 72:
        return "D1A"
    return "C1A"


def source_path(filename: str):
    working_copy = WORKING_COPY_DIR / filename
    return working_copy if working_copy.exists() else SOURCE_DIR / filename


_, metadata_records = read_records(source_path("LMS new list - master.xlsx"))
for record in metadata_records:
    record["Verticals"] = normalize_verticals(record.get("Verticals"))

metadata_ids = {value for record in metadata_records if (value := course_id(record))}

lms_by_id = {}
total_lms_rows = 0
for filename in LMS_FILES:
    _, records = read_records(SOURCE_DIR / filename)
    total_lms_rows += len(records)
    for record in records:
        identifier = course_id(record)
        if identifier in metadata_ids:
            lms_by_id[identifier] = canonical_lms_record(record)

topic_headers, topic_rows = read_records(source_path("LMS new list - Topics.xlsx"))
topics_by_course_id = defaultdict(list)
all_topic_ids = set()
total_topic_assignments = 0
matched_topic_assignments = 0
seen_assignments = set()
for column_index, raw_topic in enumerate(topic_headers):
    topic = " ".join(raw_topic.replace("\u00a0", " ").split())
    if not topic:
        continue
    vertical = topic_vertical(column_index)
    for row in topic_rows:
        value = row.get(raw_topic)
        if value in (None, ""):
            continue
        identifier = str(int(value)) if isinstance(value, float) and value.is_integer() else str(value).strip()
        all_topic_ids.add(identifier)
        total_topic_assignments += 1
        key = (identifier, topic)
        if identifier in metadata_ids and key not in seen_assignments:
            seen_assignments.add(key)
            topics_by_course_id[identifier].append({"topic": topic, "vertical": vertical})
            matched_topic_assignments += 1

content_types = Counter(
    str(record.get("Content Type")).strip()
    for record in metadata_records
    if record.get("Content Type") not in (None, "")
)

payload = {
    "generatedAt": "2026-07-31T20:00:00.000Z",
    "metadataRows": metadata_records,
    "lmsRowsByCourseId": lms_by_id,
    "topicsByCourseId": topics_by_course_id,
    "stats": {
        "totalLmsRows": total_lms_rows,
        "metadataRows": len(metadata_records),
        "matchedCourses": len(lms_by_id),
        "metadataOnlyCourses": len(metadata_ids - set(lms_by_id)),
        "topicColumns": len(topic_headers),
        "topicAssignments": total_topic_assignments,
        "matchedTopicAssignments": matched_topic_assignments,
        "matchedTopicCourseIds": len(topics_by_course_id),
        "unknownTopicCourseIds": len(all_topic_ids - metadata_ids),
        "contentTypes": dict(content_types),
    },
}

course_index = []
for record in metadata_records:
    identifier = course_id(record)
    if not identifier:
        continue
    vertical_values = [
        value.strip()
        for value in str(record.get("Verticals") or "").replace(";", ",").split(",")
        if value.strip()
    ]
    inferred_vertical = next(
        (
            assignment["vertical"]
            for assignment in topics_by_course_id.get(identifier, [])
            if assignment.get("vertical")
        ),
        "P1A",
    )
    course_index.append(
        {
            "id": identifier,
            "title": str(record.get("Course Title") or f"Course {identifier}").strip(),
            "courseCode": str(record.get("Project Code") or f"CT-{identifier}").strip(),
            "primaryVertical": vertical_values[0] if vertical_values else inferred_vertical,
        }
    )

OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
OUTPUT_PATH.write_text(
    json.dumps(payload, ensure_ascii=False, separators=(",", ":")),
    encoding="utf-8",
)
INDEX_OUTPUT_PATH.write_text(
    json.dumps(course_index, ensure_ascii=False, separators=(",", ":")),
    encoding="utf-8",
)
print(
    json.dumps(
        {
            "output": str(OUTPUT_PATH),
            "indexOutput": str(INDEX_OUTPUT_PATH),
            **payload["stats"],
        },
        indent=2,
    )
)
